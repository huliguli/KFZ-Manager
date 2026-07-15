"""Erzeugt den HSN/TSN-Seed aus der KBA-Statistik FZ 6.

Hintergrund: Die Zulassungsbescheinigung Teil I nennt in Feld 2.1 die
Herstellerschlüsselnummer (HSN) und in Feld 2.2 die Typschlüsselnummer (TSN).
Mit dem KBA-Bestandsverzeichnis lassen sich daraus Hersteller und Handelsname
auflösen — zwei Zahlen abtippen statt drei Dropdowns durchklicken.

Quelle/Lizenz (im Impressum der Original-XLSX wörtlich nachlesbar):
    © Kraftfahrt-Bundesamt, Flensburg — FZ 6 „Bestand an Kraftfahrzeugen und
    Kraftfahrzeuganhängern nach Herstellern und Typen"
    Datenlizenz Deutschland – Namensnennung – Version 2.0
    https://www.govdata.de/dl-de/by-2-0

Die Lizenz erlaubt auch kommerzielle Weiterverbreitung und Veränderung, sofern
der Quellenvermerk erhalten bleibt (siehe data/katalog/LICENSE-DATA.md) — genau
dafür schreibt dieses Skript den Vermerk mit in die Seed-Datei.

Aufruf (nur bei der jährlichen KBA-Aktualisierung nötig, nicht bei jedem Build):
    python tools/build_kba_seed.py <fz6.xlsx>
Ergebnis: src/database/kba_seed.csv.gz  (klein genug zum Mitliefern)
"""

from __future__ import annotations

import csv
import gzip
import io
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "src" / "database" / "kba_seed.csv.gz"

# Spaltenpositionen im Blatt „FZ 6.1" (1-basiert, Stand FZ 6 2025):
#   B = HSN | C = Herstellerklartext | D = TSN | E = Handelsname | F = Anzahl
SHEET = "FZ 6.1"
COL_HSN, COL_HERSTELLER, COL_TSN, COL_NAME = 2, 3, 4, 5

QUELLE = ("Kraftfahrt-Bundesamt (KBA), FZ 6 Bestand an Kraftfahrzeugen und "
          "Kraftfahrzeuganhaengern nach Herstellern und Typen; "
          "Datenlizenz Deutschland - Namensnennung - Version 2.0 "
          "(https://www.govdata.de/dl-de/by-2-0)")


def _rows(xlsx: Path):
    """Alle (hsn, tsn, hersteller, handelsname)-Zeilen des Datenblatts.

    Die Kopfzeilen (Deckblatt-Text, Überschriften) werden nicht per fester
    Zeilennummer übersprungen, sondern über die Plausibilität der Werte
    erkannt: HSN ist 4-stellig numerisch, TSN 3-stellig alphanumerisch. So
    überlebt der Import ein verschobenes Layout der nächsten Ausgabe.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Blatt '{SHEET}' fehlt — Struktur der XLSX geprüft? "
                         f"Vorhanden: {wb.sheetnames}")
    seen: set[tuple[str, str]] = set()
    for row in wb[SHEET].iter_rows(min_row=1, values_only=True):
        def cell(idx: int) -> str:
            value = row[idx - 1] if len(row) >= idx else None
            return str(value).strip() if value is not None else ""

        hsn, tsn = cell(COL_HSN), cell(COL_TSN)
        hersteller, name = cell(COL_HERSTELLER), cell(COL_NAME)
        if not (len(hsn) == 4 and hsn.isdigit()):
            continue                      # Kopf-/Leerzeile
        if not (len(tsn) == 3 and tsn.isalnum()):
            continue
        if not hersteller or not name:
            continue
        if (hsn, tsn) in seen:            # HSN/TSN ist der Primärschlüssel
            continue
        seen.add((hsn, tsn))
        yield hsn, tsn, hersteller, name


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    xlsx = Path(sys.argv[1])
    if not xlsx.is_file():
        raise SystemExit(f"Datei nicht gefunden: {xlsx}")

    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(["# quelle", QUELLE])
    writer.writerow(["hsn", "tsn", "hersteller", "handelsname"])
    count = 0
    for row in _rows(xlsx):
        writer.writerow(row)
        count += 1
    if count < 10_000:
        raise SystemExit(f"Nur {count} Zeilen erkannt — Layout der XLSX "
                         f"vermutlich geändert. Import abgebrochen (fail closed).")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with gzip.open(OUT, "wt", encoding="utf-8", newline="") as fh:
        fh.write(buffer.getvalue())
    print(f"{count} HSN/TSN-Zeilen -> {OUT} ({OUT.stat().st_size / 1024:.0f} KB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
